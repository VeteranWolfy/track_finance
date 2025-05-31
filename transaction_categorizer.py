import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from datetime import datetime
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import seaborn as sns
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import pickle
import numpy as np

class TransactionCategorizer:
    def __init__(self, root):
        self.root = root
        self.root.title("Transaction Categorizer")
        self.root.geometry("1200x800")
        
        # Store existing transactions
        self.existing_transactions = []
        
        # Store figure reference
        self.fig = None
        self.canvas = None
        
        # Bind window closing event
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Predefined categories
        self.categories = {
            '1': 'Food',
            '2': 'Transportation',
            '3': 'Entertainment',
            '4': 'Bills & Utilities & Accomodation',
            '5': 'Personal Items',
            '6': 'Income',
            '7': 'Gifts',
            '8': 'Projects',
            '9': 'Holidays',
            '0': 'Other'
        }
        
        self.current_index = 0
        self.transactions = None
        self.categorized_data = []
        
        # Create main container
        self.main_container = ttk.PanedWindow(root, orient=tk.HORIZONTAL)
        self.main_container.pack(fill=tk.BOTH, expand=True)
        
        # Create left and right frames
        self.left_frame = ttk.Frame(self.main_container)
        self.right_frame = ttk.Frame(self.main_container)
        
        self.main_container.add(self.left_frame)
        self.main_container.add(self.right_frame)
        
        # Create UI elements
        self.create_widgets()
        
        # Bind keyboard events
        self.root.bind('<Key>', self.handle_keypress)
        
    def create_widgets(self):
        # Left frame - Categorization interface
        # File selection button
        self.file_button = tk.Button(self.left_frame, text="Select File", 
                                   command=self.load_file)
        self.file_button.pack(pady=10)
        
        # Transaction display frame
        self.transaction_frame = tk.Frame(self.left_frame)
        self.transaction_frame.pack(pady=20, padx=20, fill="both", expand=True)
        
        # Transaction details
        self.date_label = tk.Label(self.transaction_frame, text="", font=("Arial", 12))
        self.date_label.pack()
        
        self.description_label = tk.Label(self.transaction_frame, text="", font=("Arial", 14, "bold"))
        self.description_label.pack()
        
        self.cost_label = tk.Label(self.transaction_frame, text="", font=("Arial", 12))
        self.cost_label.pack()
        
        # Category buttons frame
        self.category_frame = tk.Frame(self.left_frame)
        self.category_frame.pack(pady=20)
        
        # Create category buttons
        for key, value in self.categories.items():
            btn = tk.Button(self.category_frame, text=f"{key}: {value}",
                          command=lambda k=key: self.categorize_transaction(k))
            btn.pack(side=tk.LEFT, padx=5)
        
        # Navigation buttons
        self.nav_frame = tk.Frame(self.left_frame)
        self.nav_frame.pack(pady=10)
        
        self.prev_button = tk.Button(self.nav_frame, text="Previous",
                                   command=self.previous_transaction)
        self.prev_button.pack(side=tk.LEFT, padx=5)
        
        self.next_button = tk.Button(self.nav_frame, text="Next",
                                   command=self.next_transaction)
        self.next_button.pack(side=tk.LEFT, padx=5)
        
        # Save button
        self.save_button = tk.Button(self.left_frame, text="Save Categorized Data",
                                   command=self.save_categorized_data)
        self.save_button.pack(pady=10)
        
        # Right frame - Analysis
        self.fig, self.ax = plt.subplots(figsize=(8, 6))
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.right_frame)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def handle_keypress(self, event):
        if event.char in self.categories:
            self.categorize_transaction(event.char)
    
    def load_existing_transactions(self, excel_file):
        """Load existing transactions from Excel file"""
        if not os.path.exists(excel_file):
            return []
        
        existing_transactions = []
        wb = openpyxl.load_workbook(excel_file)
        
        for sheet_name in wb.sheetnames:
            if sheet_name == "Dashboard":  # Skip dashboard sheet
                continue
                
            ws = wb[sheet_name]
            current_col = 1
            
            # For each category
            for category in self.categories.values():
                row = 3  # Start after headers
                
                # Read transactions for this category
                while True:
                    date_cell = ws.cell(row=row, column=current_col)
                    if not date_cell.value:
                        break
                        
                    desc_cell = ws.cell(row=row, column=current_col + 1)
                    cost_cell = ws.cell(row=row, column=current_col + 2)
                    
                    # Handle cost value
                    try:
                        cost_value = float(cost_cell.value if cost_cell.value is not None else 0)
                    except (ValueError, TypeError):
                        cost_value = 0
                    
                    # Handle date value
                    try:
                        if isinstance(date_cell.value, datetime):
                            date_value = date_cell.value.strftime('%Y-%m-%d')
                        else:
                            date_value = datetime.strptime(str(date_cell.value), '%Y-%m-%d').strftime('%Y-%m-%d')
                    except (ValueError, TypeError):
                        # Skip invalid date entries
                        row += 1
                        continue
                    
                    existing_transactions.append({
                        'date': date_value,
                        'description': str(desc_cell.value) if desc_cell.value else '',
                        'cost': cost_value,
                        'category': category
                    })
                    row += 1
                
                current_col += 3
        
        return existing_transactions

    def is_duplicate(self, transaction):
        """Check if a transaction is a duplicate"""
        for existing in self.existing_transactions:
            if (str(existing['date']) == str(transaction['date']) and
                existing['description'] == transaction['description'] and
                abs(existing['cost'] - transaction['cost']) < 0.01):  # Using small delta for float comparison
                return True
        return False

    def load_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[
                ("All supported files", "*.csv *.xlsx *.xls"),
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*")
            ]
        )
        if file_path:
            try:
                # First, ask for the Excel file to check duplicates against
                excel_path = filedialog.askopenfilename(
                    title="Select existing Excel file (Cancel if this is your first import)",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
                )
                
                if excel_path:
                    self.excel_path = excel_path
                    self.existing_transactions = self.load_existing_transactions(excel_path)
                
                # Load file based on extension
                file_extension = os.path.splitext(file_path)[1].lower()
                if file_extension in ['.xlsx', '.xls']:
                    df = pd.read_excel(file_path)
                else:  # Try CSV with different encodings
                    try:
                        df = pd.read_csv(file_path, encoding='utf-8-sig')
                    except UnicodeDecodeError:
                        try:
                            df = pd.read_csv(file_path, encoding='latin1')
                        except:
                            df = pd.read_csv(file_path, encoding='cp1252')
                
                # Clean up data
                df = df.replace([np.inf, -np.inf], np.nan)  # Replace infinite values with NaN
                df = df.fillna(0)  # Replace NaN with 0 for numerical columns
                
                # Handle different formats
                # Tesco credit card format
                if 'Amount' in df.columns and 'Merchant' in df.columns:
                    # Convert Amount to numeric, ensuring it's properly formatted
                    df['Amount'] = pd.to_numeric(df['Amount'].astype(str).str.replace('£', '').str.replace(',', ''), errors='coerce')
                    
                    # Handle Direct Debit payments - exclude them
                    direct_debit_mask = df['Merchant'].str.contains('DIRECT DEBIT PAYMENT', case=False, na=False)
                    df = df[~direct_debit_mask]
                    
                    # For remaining transactions, positive amounts should be treated as costs (negative)
                    df['cost'] = df['Amount'].apply(lambda x: -abs(x))
                    
                    # Map other columns
                    if 'Date' in df.columns:
                        df = df.rename(columns={'Date': 'date'})
                    if 'Merchant' in df.columns:
                        df = df.rename(columns={'Merchant': 'description'})
                
                # Handle other common column names
                column_mappings = {
                    'Transaction Date': 'date',
                    'Trans Date': 'date',
                    'Date': 'date',
                    'Transaction Description': 'description',
                    'Description': 'description',
                    'Details': 'description',
                    'Merchant': 'description',
                    'Amount': 'cost',
                    'Value': 'cost',
                    'Billing Amount': 'cost',
                    'Transaction Amount': 'cost'
                }
                
                df = df.rename(columns=column_mappings)
                
                # Ensure required columns exist
                required_columns = ['date', 'description', 'cost']
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")
                
                # Convert date strings to datetime with various formats
                try:
                    df['date'] = pd.to_datetime(df['date'], dayfirst=True)
                except:
                    try:
                        # Try specific format if automatic parsing fails
                        df['date'] = pd.to_datetime(df['date'], format='%d/%m/%Y')
                    except:
                        # Try other common formats
                        date_formats = ['%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%d-%m-%y']
                        for fmt in date_formats:
                            try:
                                df['date'] = pd.to_datetime(df['date'], format=fmt)
                                break
                            except:
                                continue
                
                df['date'] = df['date'].dt.strftime('%Y-%m-%d')
                
                # Clean up description
                df['description'] = df['description'].astype(str).str.strip()
                
                # Clean up cost values
                if isinstance(df['cost'].iloc[0], str):
                    # Remove currency symbols and commas, then convert to float
                    df['cost'] = df['cost'].str.replace('£', '').str.replace(',', '')
                df['cost'] = pd.to_numeric(df['cost'], errors='coerce').fillna(0)
                
                # Filter duplicates
                filtered_transactions = []
                duplicates_count = 0
                
                for _, row in df.iterrows():
                    transaction = {
                        'date': row['date'],
                        'description': row['description'],
                        'cost': float(row['cost'])
                    }
                    
                    if not self.is_duplicate(transaction):
                        filtered_transactions.append(transaction)
                    else:
                        duplicates_count += 1
                
                if duplicates_count > 0:
                    messagebox.showinfo("Duplicates Found", 
                                      f"{duplicates_count} duplicate transactions were found and skipped.")
                
                if not filtered_transactions:
                    messagebox.showinfo("No New Transactions", 
                                      "All transactions in the file are duplicates.")
                    return
                
                self.transactions = pd.DataFrame(filtered_transactions)
                self.current_index = 0
                self.categorized_data = []
                self.display_current_transaction()
                self.update_pie_chart()
                
            except Exception as e:
                messagebox.showerror("Error", f"Error loading file: {str(e)}")
                # Print more detailed error information
                import traceback
                print(traceback.format_exc())

    def display_current_transaction(self):
        if self.transactions is not None and self.current_index < len(self.transactions):
            transaction = self.transactions.iloc[self.current_index]
            self.date_label.config(text=f"Date: {transaction['date']}")
            self.description_label.config(text=f"Description: {transaction['description']}")
            self.cost_label.config(text=f"Amount: £{transaction['cost']:.2f}")
            
            # Update navigation buttons state
            self.prev_button.config(state=tk.NORMAL if self.current_index > 0 else tk.DISABLED)
            self.next_button.config(state=tk.NORMAL if self.current_index < len(self.transactions) - 1 else tk.DISABLED)
    
    def categorize_transaction(self, category_key):
        if self.transactions is not None and self.current_index < len(self.transactions):
            transaction = self.transactions.iloc[self.current_index]
            category = self.categories.get(category_key, "Other")  # Default to "Other" if key not found
            self.categorized_data.append({
                'date': transaction['date'],
                'description': transaction['description'],
                'cost': transaction['cost'],
                'category': category
            })
            self.next_transaction()
            self.update_pie_chart()
    
    def next_transaction(self):
        if self.current_index < len(self.transactions) - 1:
            self.current_index += 1
            self.display_current_transaction()
    
    def previous_transaction(self):
        if self.current_index > 0:
            self.current_index -= 1
            self.display_current_transaction()
    
    def save_categorized_data(self):
        if not self.categorized_data:
            messagebox.showwarning("Warning", "No categorized data to save!")
            return
        
        if not hasattr(self, 'excel_path'):
            self.excel_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if not self.excel_path:
                return
        
        try:
            # Convert new data to DataFrame
            df = pd.DataFrame(self.categorized_data)
            df['date'] = pd.to_datetime(df['date'])
            df = df.sort_values('date')
            
            # Create new workbook or load existing one
            if os.path.exists(self.excel_path):
                try:
                    wb = openpyxl.load_workbook(self.excel_path)
                except:
                    # If file is corrupted, create new workbook
                    wb = openpyxl.Workbook()
            else:
                wb = openpyxl.Workbook()
            
            # Remove default sheet if it exists and no data has been added to it
            if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
                wb.remove(wb['Sheet'])
            if 'Sheet1' in wb.sheetnames and len(wb.sheetnames) > 1:
                wb.remove(wb['Sheet1'])
            
            # Process new transactions by month
            # Group by year and month
            df['year_month'] = df['date'].dt.strftime('%Y-%m')
            for year_month in df['year_month'].unique():
                month_data = df[df['year_month'] == year_month]
                sheet_name = month_data['date'].dt.strftime('%B %Y').iloc[0]
                
                # Get or create worksheet
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(sheet_name)
                    self.setup_worksheet_headers(ws)
                
                # Find last row for each category and append data
                for category in self.categories.values():
                    category_data = month_data[month_data['category'] == category]
                    if category_data.empty:
                        continue
                    
                    category_col = list(self.categories.values()).index(category) * 3 + 1
                    row = 3
                    
                    # Find last row in this category
                    while ws.cell(row=row, column=category_col).value is not None:
                        row += 1
                    
                    # Append new data
                    for _, transaction in category_data.iterrows():
                        ws.cell(row=row, column=category_col, value=transaction['date'].strftime('%Y-%m-%d'))
                        ws.cell(row=row, column=category_col+1, value=transaction['description'])
                        cost_cell = ws.cell(row=row, column=category_col+2, value=transaction['cost'])
                        cost_cell.number_format = '£#,##0.00'
                        
                        # Add borders
                        for col in range(category_col, category_col+3):
                            cell = ws.cell(row=row, column=col)
                            cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                        
                        # Add thick right border
                        ws.cell(row=row, column=category_col+2).border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thick'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        row += 1
            
            # Save workbook
            wb.save(self.excel_path)
            messagebox.showinfo("Success", "Data appended successfully!")
            
            # Update existing transactions list
            self.existing_transactions.extend(self.categorized_data)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error saving data: {str(e)}")

    def setup_worksheet_headers(self, ws):
        """Set up headers for a new worksheet"""
        # Define styles
        header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        header_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set up column headers
        current_col = 1
        for category in self.categories.values():
            # Category header
            ws.cell(row=1, column=current_col, value=category)
            ws.cell(row=1, column=current_col+1, value="")
            ws.cell(row=1, column=current_col+2, value="")
            
            # Subheaders
            ws.cell(row=2, column=current_col, value="Date")
            ws.cell(row=2, column=current_col+1, value="Description")
            ws.cell(row=2, column=current_col+2, value="Cost")
            
            # Style headers
            for col in range(current_col, current_col+3):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell = ws.cell(row=2, column=col)
                cell.font = header_font
                cell.border = border
            
            # Set currency format for cost column
            for row in range(3, 1000):  # Pre-format a reasonable number of rows
                ws.cell(row=row, column=current_col+2).number_format = '£#,##0.00'
            
            # Add right border to last column of each category
            ws.cell(row=1, column=current_col+2).border = Border(
                left=Side(style='thin'),
                right=Side(style='thick'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            ws.cell(row=2, column=current_col+2).border = Border(
                left=Side(style='thin'),
                right=Side(style='thick'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            current_col += 3
        
        # Adjust column widths
        for col in range(1, current_col):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def update_pie_chart(self):
        if not self.categorized_data:
            return
            
        df = pd.DataFrame(self.categorized_data)
        df['date'] = pd.to_datetime(df['date'])
        
        # Clear previous figure
        if self.fig is not None:
            plt.close(self.fig)
        if self.canvas is not None:
            self.canvas.get_tk_widget().destroy()
        
        # Create new figure
        self.fig, self.ax = plt.subplots(figsize=(8, 6))
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.right_frame)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        # Trends Analysis
        df['period'] = df['date'].dt.strftime('%Y-%m')
        
        # Create trends plot
        sns.barplot(x='period', y='cost', hue='category', data=df, ax=self.ax)
        self.ax.set_title('Monthly Income/Expenses')
        self.ax.set_xlabel('Month')
        self.ax.set_ylabel('Amount (£)')
        plt.xticks(rotation=45)
        plt.tight_layout()
        
        self.canvas.draw()

    def create_dashboard(self, wb):
        """Create or update the dashboard sheet with monthly summaries"""
        sheet_name = "Dashboard"
        
        # Check if dashboard exists
        if sheet_name in wb.sheetnames:
            # Get existing dashboard
            ws = wb[sheet_name]
            
            # Only clear the data cells, not the entire sheet
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell.value, (int, float, str)):  # Only clear data cells
                        cell.value = None
        else:
            # Create new dashboard sheet
            ws = wb.create_sheet(sheet_name, 0)  # Add dashboard as first sheet
        
        # Define styles
        header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        header_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Collect all transaction data
        all_transactions = []
        for sheet_name in wb.sheetnames:
            if sheet_name == "Dashboard":
                continue
            
            sheet = wb[sheet_name]
            current_col = 1
            
            for category in self.categories.values():
                row = 3
                while True:
                    date_cell = sheet.cell(row=row, column=current_col)
                    if not date_cell.value:
                        break
                    
                    cost_cell = sheet.cell(row=row, column=current_col + 2)
                    
                    if isinstance(date_cell.value, str):
                        try:
                            date_value = datetime.strptime(date_cell.value, '%Y-%m-%d')
                        except ValueError:
                            date_value = date_cell.value
                    else:
                        date_value = date_cell.value
                    
                    all_transactions.append({
                        'date': date_value,
                        'category': category,
                        'cost': float(cost_cell.value) if cost_cell.value is not None else 0
                    })
                    row += 1
                current_col += 3
        
        if not all_transactions:
            return
        
        # Convert to DataFrame for easier analysis
        df = pd.DataFrame(all_transactions)
        df['date'] = pd.to_datetime(df['date'])
        df['month'] = df['date'].dt.strftime('%B %Y')
        
        # Create monthly summary
        # Headers
        ws.cell(row=1, column=1, value="Month")
        current_col = 2
        for category in self.categories.values():
            ws.cell(row=1, column=current_col, value=category)
            current_col += 1
        ws.cell(row=1, column=current_col, value="Monthly Total")
        
        # Style headers
        for col in range(1, current_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Monthly data
        current_row = 2
        monthly_summary = df.pivot_table(
            index='month',
            columns='category',
            values='cost',
            aggfunc='sum',
            fill_value=0
        )
        
        # Sort months chronologically
        monthly_summary.index = pd.to_datetime(monthly_summary.index, format='%B %Y')
        monthly_summary = monthly_summary.sort_index()
        monthly_summary.index = monthly_summary.index.strftime('%B %Y')
        
        for month in monthly_summary.index:
            ws.cell(row=current_row, column=1, value=month)
            current_col = 2
            row_total = 0
            
            for category in self.categories.values():
                value = monthly_summary.loc[month, category] if category in monthly_summary.columns else 0
                ws.cell(row=current_row, column=current_col, value=value)
                row_total += value
                current_col += 1
            
            ws.cell(row=current_row, column=current_col, value=row_total)
            
            # Add borders to row
            for col in range(1, current_col + 1):
                ws.cell(row=current_row, column=col).border = border
            
            current_row += 1
        
        # Add yearly totals
        ws.cell(row=current_row, column=1, value="Year Total")
        ws.cell(row=current_row, column=1).font = header_font
        current_col = 2
        year_total = 0
        
        for category in self.categories.values():
            total = monthly_summary[category].sum() if category in monthly_summary.columns else 0
            ws.cell(row=current_row, column=current_col, value=total)
            year_total += total
            current_col += 1
        
        ws.cell(row=current_row, column=current_col, value=year_total)
        
        # Style yearly totals row
        for col in range(1, current_col + 1):
            cell = ws.cell(row=current_row, column=col)
            cell.font = header_font
            cell.border = border
        
        # Format all numbers as currency
        for row in range(2, current_row + 1):
            for col in range(2, current_col + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '£#,##0.00'
        
        # Adjust column widths
        for col in range(1, current_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def on_closing(self):
        """Handle window closing event"""
        if self.fig is not None:
            plt.close(self.fig)
        if hasattr(self, 'canvas') and self.canvas is not None:
            self.canvas.get_tk_widget().destroy()
        self.root.quit()
        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = TransactionCategorizer(root)
    root.mainloop() 